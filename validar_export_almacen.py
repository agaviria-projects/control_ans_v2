"""
------------------------------------------------------------
CONTROL DE ALMACÉN – CRUCE FÉNIX (EPM) vs ELITE (Planilla Consumos)
Versión 3.2 - 2025
Autor: Héctor + IA
------------------------------------------------------------
Descripción:
  Cruza los registros de EPM (archivo FÉNIX) con la planilla
  interna de ELITE (archivo .xlsm), comparando PEDIDOS, CÓDIGOS
  y opcionalmente MANO DE OBRA. Calcula diferencias y genera:
    - Hoja CONTROL_ALMACEN con detalle fila a fila
    - Hoja RESUMEN con totales por estado
------------------------------------------------------------
"""

# ============================================================
# 1. LIBRERÍAS
# ============================================================
from pathlib import Path
import pandas as pd
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================
# 2. CONFIGURACIÓN DE RUTAS
# ============================================================
base = Path(__file__).resolve().parent
ruta_fenix = base / "data_raw" / "Digitacion Fenix.xlsx"
ruta_elite = base / "data_raw" / "Planilla consumos.xlsx"
ruta_salida = base / "data_clean" / "CONTROL_ALMACEN.xlsx"

print("------------------------------------------------------------")
print("🚀 INICIANDO CRUCE FÉNIX vs ELITE (v3.2)...")
print("------------------------------------------------------------")

# ============================================================
# 3. CARGA DE DATOS
# ============================================================

columnas_fenix = [
    "pedido", "subz", "municipio", "contrato", "acta", "actividad",
    "fecha_estado", "pagina", "urbrur", "tipre", "red_interna",
    "tipo_operacion", "tipo", "cobro", "suminis", "item_cont",
    "item_res", "cantidad", "vlr_cliente", "valor_costo"
]

# --- FÉNIX ---
try:
    df_fenix = pd.read_excel(ruta_fenix, dtype=str)
    df_fenix.columns = df_fenix.columns.str.lower().str.strip()
    df_fenix = df_fenix[[c for c in columnas_fenix if c in df_fenix.columns]]
    df_fenix["cantidad_fenix"] = pd.to_numeric(df_fenix["cantidad"], errors="coerce").fillna(0)
    if "mano_obra" not in df_fenix.columns:
        df_fenix["mano_obra"] = None
except Exception as e:
    raise SystemExit(f"❌ Error al leer FÉNIX: {e}")

# --- ELITE ---
try:
    print("🔎 Buscando hoja con encabezado 'CANTIDAD' y limpiando filas previas...")

    temp = pd.read_excel(ruta_elite, sheet_name=None, dtype=str, header=None)
    hoja_correcta, fila_header = None, None

    hojas_excluir = ["pasos", "resumen", "hoja1", "macro", "configuracion"]

    for hoja, df_temp in temp.items():
        if hoja.lower() in hojas_excluir:
            continue

        # Buscar fila con la palabra 'cantidad' (encabezado real)
        for i, fila in df_temp.iterrows():
            fila_texto = " ".join(str(x).lower() for x in fila.values if pd.notna(x))
            if "cantidad" in fila_texto:
                hoja_correcta, fila_header = hoja, i
                break
        if hoja_correcta:
            break

    if hoja_correcta is None:
        raise Exception("No se encontró ninguna hoja con encabezado 'CANTIDAD'.")

    print(f"📍 Hoja detectada: {hoja_correcta} | Encabezado en fila {fila_header + 1}")

    # 🔹 Leer solo desde el encabezado hacia abajo (evita filas vacías previas)
    df_elite = pd.read_excel(
        ruta_elite,
        sheet_name=hoja_correcta,
        dtype=str,
        skiprows=fila_header  # <-- lee desde esa fila hacia abajo
    )

    # 🔹 Limpiar encabezados
    df_elite.columns = (
        df_elite.columns
        .map(str)
        .str.lower()
        .str.strip()
        .str.replace(r"unnamed.*", "", regex=True)
    )

    # 🔹 Quitar columnas completamente vacías
    df_elite = df_elite.dropna(axis=1, how="all")

    # 🔹 Conservar columnas relevantes
    posibles_cols = ["#pedido", "pedido", "codigo", "cantidad", "descripción", "descripcion", "unidad", "tecnico"]
    df_elite = df_elite[[c for c in df_elite.columns if any(p in c for p in posibles_cols)]]

    # 🔹 Renombrar columnas estándar
    df_elite.rename(columns={
        "#pedido": "pedido",
        "codigo": "codigo",
        "cantidad": "cantidad_elite"
    }, inplace=True)

    # 🔹 Filtrar filas vacías o corruptas (pedidos falsos)
    df_elite["pedido"] = df_elite["pedido"].astype(str).str.strip()
    df_elite = df_elite[
        df_elite["pedido"].notna() &
        df_elite["pedido"].str.match(r"^\d{8,}$", na=False)
    ]

    # 🔹 Limpiar códigos válidos (6 dígitos)
    df_elite = df_elite[df_elite["codigo"].astype(str).str.match(r"^\d{6}$", na=False)]

    # 🔹 Convertir cantidades a número
    df_elite["cantidad_elite"] = pd.to_numeric(df_elite["cantidad_elite"], errors="coerce").fillna(0)

    # 🔹 Agregar columna mano_obra si no existe
    if "mano_obra" not in df_elite.columns:
        df_elite["mano_obra"] = None

    print(f"✅ Planilla Consumos lista para cruce: {len(df_elite)} registros limpios.")

except Exception as e:
    raise SystemExit(f"❌ Error al leer Planilla Consumos: {e}")



print("✅ Archivos cargados correctamente.")
time.sleep(0.5)

# ============================================================
# 4. CRUCE PRINCIPAL (FÉNIX vs ELITE)
# ============================================================
print("⚙️ Ejecutando cruce principal FÉNIX vs ELITE...")

df_fenix.rename(columns={"item_res": "codigo"}, inplace=True)

# Validar columnas clave
for col in ["pedido", "codigo"]:
    if col not in df_fenix.columns:
        df_fenix[col] = None
    if col not in df_elite.columns:
        df_elite[col] = None

# Filtrar códigos válidos (solo 6 dígitos)
df_elite = df_elite[df_elite["codigo"].astype(str).str.match(r"^\d{6}$", na=False)]

# ============================================================
# 4.1. Normalizar códigos base y complementarios antes del merge
# ============================================================

# 🔹 Definir equivalencias base ↔ complemento
equivalencias = {
    "200492A": "200492",
    "200384A": "200384"
}

# 🔹 Crear columna auxiliar con el código base normalizado
df_fenix["codigo_equiv"] = df_fenix["codigo"].replace(equivalencias)
df_elite["codigo_equiv"] = df_elite["codigo"].replace(equivalencias)

# 🔹 Agregar columna 'origen' antes del merge (evita KeyError)
df_fenix["origen"] = "FENIX"
df_elite["origen"] = "ELITE"

# 🔹 Merge extendido usando el código normalizado
df_full = pd.merge(
    df_fenix,
    df_elite[["pedido", "codigo_equiv", "cantidad_elite", "origen"]],
    left_on=["pedido", "codigo_equiv"],
    right_on=["pedido", "codigo_equiv"],
    how="outer",
    indicator=True
)

# 🔹 Renombrar para mantener compatibilidad con el resto del código
df_full.rename(columns={"codigo_equiv": "codigo"}, inplace=True)

# ============================================================
# 🔧 Limpieza de duplicados tras merge extendido
# ============================================================
# Eliminar columnas duplicadas (mantiene solo la primera aparición)
df_full = df_full.loc[:, ~df_full.columns.duplicated()].copy()

# En caso de que queden versiones 'codigo_x' o 'codigo_y', unificarlas
if "codigo_x" in df_full.columns:
    df_full["codigo"] = df_full["codigo_x"].combine_first(df_full.get("codigo_y"))
    df_full.drop(columns=["codigo_x", "codigo_y"], errors="ignore", inplace=True)


# # ============================================================
# # 5. CRUCE COMPLETO PARA DETECTAR COINCIDENCIAS Y FALTANTES
# # ============================================================
# df_fenix["origen"] = "FENIX"
# df_elite["origen"] = "ELITE"

# # Cruce completo (outer join)
# df_full = pd.merge(
#     df_fenix,
#     df_elite[["pedido", "codigo", "cantidad_elite", "origen"]],
#     on=["pedido", "codigo"],
#     how="outer",
#     indicator=True
# )

# ============================================================
# 6. GENERAR SUBCONJUNTOS
# ============================================================
# Coincidencias reales (ambos archivos)
df_merge = df_full[df_full["_merge"] == "both"].copy()

# Sin cruce (solo FENIX o solo ELITE)
df_nocruce = df_full[df_full["_merge"] != "both"].copy()
df_nocruce["origen"] = df_nocruce["_merge"].replace({
    "left_only": "Solo en FENIX",
    "right_only": "Solo en ELITE"
})

# ============================================================
# 6.1. REGLA ESPECIAL – Mantener códigos complementarios válidos
# ============================================================
# No enviar a NO_COINCIDEN el código 200492A (ni sus pares)
codigos_validos = ["200492A"]

# Sacar estos registros de df_nocruce y mantenerlos en df_merge
df_extra_validos = df_nocruce[df_nocruce["codigo"].isin(codigos_validos)].copy()
if not df_extra_validos.empty:
    print(f"🧩 Registros especiales mantenidos en CONTROL_ALMACEN: {len(df_extra_validos)}")
    df_extra_validos["estado"] = "OK – Material Complementario"
    df_extra_validos["diferencia"] = 0
    df_merge = pd.concat([df_merge, df_extra_validos], ignore_index=True)

    # Quitar estos del listado de no coincidentes
    df_nocruce = df_nocruce[~df_nocruce["codigo"].isin(codigos_validos)]


# ============================================================
# 7. CÁLCULO DE DIFERENCIA Y ESTADO
# ============================================================
df_merge["cantidad_fenix"] = pd.to_numeric(df_merge.get("cantidad", 0), errors="coerce").fillna(0)
df_merge["cantidad_elite"] = pd.to_numeric(df_merge.get("cantidad_elite", 0), errors="coerce").fillna(0)
df_merge["diferencia"] = df_merge["cantidad_fenix"] - df_merge["cantidad_elite"]

def evaluar(row):
    if row["diferencia"] == 0:
        return "OK"
    elif row["diferencia"] > 0:
        return "FALTANTE EN ELITE"
    else:
        return "EXCESO EN ELITE"

df_merge["estado"] = df_merge.apply(evaluar, axis=1)
# ============================================================
# 7.1. AJUSTE DE MATERIALES COMPLEMENTARIOS (mantiene ambos códigos visibles)
# ============================================================

# 🔹 Diccionario base ↔ complemento (Se puede ampliar sin modificar lógica)
complementos = {
    "200492": "200492A",
    "200384": "200384A"
}

ajustes_realizados = 0

# 🔹 1. Ajuste en df_merge (CONTROL_ALMACEN)
for pedido in df_merge["pedido"].unique():
    for base, comp in complementos.items():
        # Filtrar registros del mismo pedido con el código base o su complemento
        grupo = df_merge[
            (df_merge["pedido"] == pedido)
            & (df_merge["codigo"].isin([base, comp]))
        ]

        if not grupo.empty:
            total_fenix = grupo["cantidad_fenix"].sum()
            total_elite = grupo["cantidad_elite"].sum()

            # Si Elite tiene igual o más cantidad → marcar ambos como complementarios
            if total_elite >= total_fenix and total_fenix > 0:
                df_merge.loc[
                    (df_merge["pedido"] == pedido)
                    & (df_merge["codigo"].isin([base, comp])),
                    ["estado", "diferencia"]
                ] = ["OK – Material Complementario", 0]
                ajustes_realizados += 1

print(f"🔧 Ajustes aplicados (manteniendo ambos códigos): {ajustes_realizados}")

# 🔹 2. Ajuste en df_nocruce (NO_COINCIDEN)
if not df_nocruce.empty:
    registros_ajustados = 0
    for base, comp in complementos.items():
        df_nocruce = df_nocruce[
            ~(
                (df_nocruce["codigo"].isin([base, comp]))
                & (df_nocruce["pedido"].isin(df_merge["pedido"].unique()))
            )
        ]
        registros_ajustados += 1
    print(f"🧩 Registros eliminados de NO_COINCIDEN por complementarios: {registros_ajustados}")


# ============================================================
# 8. ORGANIZAR COLUMNAS FINALES
# ============================================================
columnas_fenix = [
    "pedido", "subz", "municipio", "contrato", "acta",
    "actividad", "fecha_estado", "pagina", "urbrur", "tipre",
    "red_interna", "tipo_operacion", "tipo", "cobro", "suminis",
    "item_cont", "codigo", "cantidad", "vlr_cliente", "valor_costo"
]

# Cambiamos el nombre de la columna "estado" a "status" antes del orden
if "estado" in df_merge.columns:
    df_merge.rename(columns={"estado": "status"}, inplace=True)

columnas_finales = columnas_fenix + ["cantidad_elite", "diferencia", "status"]
df_merge = df_merge[[c for c in columnas_finales if c in df_merge.columns]]

# Para hoja NO_COINCIDEN
columnas_nocruce = ["pedido", "codigo", "cantidad", "cantidad_elite", "origen"]
df_nocruce = df_nocruce[[c for c in columnas_nocruce if c in df_nocruce.columns]]

# ============================================================
# 9. CREAR RESUMEN
# ============================================================
# Agrupamos por la nueva columna "status" en lugar de "estado_final"
resumen = (
    df_merge.groupby("status", dropna=False)
    .size()
    .reset_index(name="total")
    .sort_values(by="status")
)
total_registros = len(df_merge)
resumen.loc[len(resumen)] = ["TOTAL GENERAL", total_registros]

# Cambiamos nombre de la columna del resumen a "estado_final"
resumen.rename(columns={"status": "estado_final"}, inplace=True)

# ============================================================
# 🔹 LIMPIEZA DE PEDIDOS (evita falsos pedidos 1, 2, 3…)
# ============================================================
if "pedido" in df_elite.columns:
    # Normalizar y eliminar filas sin pedido válido
    df_elite["pedido"] = (
        df_elite["pedido"]
        .astype(str)
        .str.strip()
        .replace({"nan": None, "": None})
    )

    # Conservar solo filas con pedidos numéricos reales de 8 dígitos o más
    df_elite = df_elite[
        df_elite["pedido"].notna() &
        df_elite["pedido"].str.match(r"^\d{8,}$", na=False)
    ]

    # Eliminar filas vacías restantes
    df_elite = df_elite.dropna(subset=["pedido"])


# ============================================================
# 10. EXPORTAR A EXCEL (manejo de archivo abierto)
# ============================================================
ruta_salida.parent.mkdir(parents=True, exist_ok=True)

try:
    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df_merge.to_excel(writer, index=False, sheet_name="CONTROL_ALMACEN")
        resumen.to_excel(writer, index=False, sheet_name="RESUMEN")
        df_nocruce.to_excel(writer, index=False, sheet_name="NO_COINCIDEN")

    print("💾 Exportando archivo con hoja de control de pendientes...")

except PermissionError:
    print("⚠️ No se puede guardar el archivo porque está abierto en Excel.")
    print("🧩 Por favor, cierre 'CONTROL_ALMACEN.xlsx' y ejecute nuevamente el script.")
    import sys
    sys.exit(1)

except Exception as e:
    print(f"❌ Error inesperado al exportar a Excel: {e}")
    import sys
    sys.exit(1)


# ============================================================
# 🔹 NORMALIZAR TIPOS DE DATOS (evita "Recuento" en Excel)
# ============================================================
cols_numericas = ["cantidad", "cantidad_elite", "vlr_cliente", "valor_costo", "diferencia"]

for col in cols_numericas:
    if col in df_merge.columns:
        df_merge[col] = (
            pd.to_numeric(df_merge[col], errors="coerce")
            .fillna(0)
            .astype(float)
        )
# ============================================================
# 🔹 Normalizar tipos numéricos también en NO_COINCIDEN
# ============================================================
cols_numericas_nc = ["cantidad", "cantidad_elite"]

for col in cols_numericas_nc:
    if col in df_nocruce.columns:
        df_nocruce[col] = (
            pd.to_numeric(df_nocruce[col], errors="coerce")
            .fillna(0)
            .astype(float)
        )

with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
    df_merge.to_excel(writer, index=False, sheet_name="CONTROL_ALMACEN")
    resumen.to_excel(writer, index=False, sheet_name="RESUMEN")
    df_nocruce.to_excel(writer, index=False, sheet_name="NO_COINCIDEN")

print("💾 Exportando archivo con hoja de control de pendientes...")

# Asegurar que las columnas numéricas estén en formato numérico real
for col in ["cantidad", "cantidad_elite", "vlr_cliente", "valor_costo", "diferencia"]:
    if col in df_merge.columns:
        df_merge[col] = pd.to_numeric(df_merge[col], errors="coerce").fillna(0)

# ============================================================
# 11. FORMATO VISUAL LIMPIO
# ============================================================
wb = load_workbook(ruta_salida)

def formato_hoja(ws):
    from openpyxl.styles import PatternFill, Font, Alignment

    max_row, max_col = ws.max_row, ws.max_column

    font_encabezado = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # 🎨 Paleta de colores
    colores = {
        "default": "004C99",      # azul (FENIX)
        "elite": "000000",        # morado (ELITE)
        "diferencia": "000000",   # naranja (comparativo)
        "status": "000000",       # verde (resultado)
    }

    # 🔹 Colorear encabezados según tipo
    for idx, cell in enumerate(ws[1], 1):
        header = str(cell.value).lower().strip()
        color = colores["default"]  # por defecto azul FENIX

        if "elite" in header:
            color = colores["elite"]
        elif "diferencia" in header:
            color = colores["diferencia"]
        elif header == "status":  # evitar confusión con fecha_estado
            color = colores["status"]

        cell.fill = PatternFill("solid", start_color=color)
        cell.font = font_encabezado
        cell.alignment = align_center

    # 🔹 Alinear celdas del cuerpo
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for c in row:
            c.alignment = align_center


# === CONTROL_ALMACEN ===
ws = wb["CONTROL_ALMACEN"]
formato_hoja(ws)

# Aplicar semáforo sobre columna STATUS
col_status = None
for idx, cell in enumerate(ws[1], 1):
    if str(cell.value).lower().strip() == "status":
        col_status = idx
        break

if col_status:
    for i in range(2, ws.max_row + 1):
        c = ws.cell(row=i, column=col_status)
        text = str(c.value).upper()
        if "OK" in text:
            c.fill = PatternFill("solid", start_color="00B050")
            c.font = Font(color="FFFFFF", bold=True)
        elif "FALTANTE" in text:
            c.fill = PatternFill("solid", start_color="FFD966")
            c.font = Font(color="000000", bold=True)
        elif "EXCESO" in text:
            c.fill = PatternFill("solid", start_color="C00000")
            c.font = Font(color="FFFFFF", bold=True)

# === RESUMEN ===
ws_resumen = wb["RESUMEN"]
formato_hoja(ws_resumen)

# === NO_COINCIDEN ===
ws_nc = wb["NO_COINCIDEN"]
formato_hoja(ws_nc)

for i in range(2, ws_nc.max_row + 1):
    c = ws_nc.cell(row=i, column=ws_nc.max_column)
    if "ELITE" in str(c.value).upper():
        c.fill = PatternFill("solid", start_color="C00000")
        c.font = Font(color="FFFFFF", bold=True)
    elif "FENIX" in str(c.value).upper():
        c.fill = PatternFill("solid", start_color="1F4E78")
        c.font = Font(color="FFFFFF", bold=True)

wb.save(ruta_salida)
wb.close()

print("✅ CRUCE FINALIZADO CON ÉXITO (v3.7 con colores de encabezado).")
print(f"📁 Archivo generado: {ruta_salida}")
print("------------------------------------------------------------")


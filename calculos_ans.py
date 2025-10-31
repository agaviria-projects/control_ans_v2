"""
------------------------------------------------------------
CÁLCULOS ANS – Proyecto Control_ANS_FENIX
------------------------------------------------------------
Autor: Héctor + IA (2025)
------------------------------------------------------------
Descripción:
- Lee el archivo limpio (FENIX_CLEAN.xlsx)
- Calcula días pactados, fecha límite ANS, estado y métricas.
- Excluye sábados, domingos y festivos.
- Mantiene hora/minuto del inicio.
- Exporta a FENIX_ANS.xlsx con hoja RESUMEN.
------------------------------------------------------------
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

# ------------------------------------------------------------
# CONFIGURACIÓN DE RUTAS
# ------------------------------------------------------------
base_path = Path(__file__).resolve().parent.parent
ruta_input = base_path / "Control_ANS" / "data_clean" / "FENIX_CLEAN.xlsx"
ruta_output = base_path / "Control_ANS" / "data_clean" / "FENIX_ANS.xlsx"

# ------------------------------------------------------------
# CONFIGURACIÓN DE CALENDARIO
# ------------------------------------------------------------
WEEKMASK = "1111100"  # lunes a viernes

FESTIVOS = np.array([
    "2025-01-01","2025-01-06","2025-03-24","2025-04-17","2025-04-18",
    "2025-05-01","2025-05-26","2025-06-16","2025-06-23","2025-07-07",
    "2025-08-07","2025-08-18","2025-10-13","2025-11-03","2025-11-17",
    "2025-12-08","2025-12-25",
    "2026-01-01","2026-01-12","2026-03-23","2026-04-02","2026-04-03",
    "2026-05-01","2026-05-18","2026-06-08","2026-06-15","2026-06-29",
    "2026-07-20","2026-08-07","2026-08-17","2026-10-12","2026-11-02",
    "2026-11-16","2026-12-08","2026-12-25"
], dtype="datetime64[D]")

# ------------------------------------------------------------
# FUNCIÓN: FECHA LÍMITE SEGÚN LÓGICA FÉNIX
# ------------------------------------------------------------
def add_business_days_keep_time(start_dt, n_days):
    if pd.isna(start_dt) or n_days <= 0:
        return pd.NaT

    date_part = np.datetime64(start_dt.date())
    time_part = start_dt.time()

    # Día no hábil → primer hábil siguiente
    if not np.is_busday(date_part, weekmask=WEEKMASK, holidays=FESTIVOS):
        primer_habil = np.busday_offset(date_part, 0, roll="forward",
                                        weekmask=WEEKMASK, holidays=FESTIVOS)
        limite = np.busday_offset(primer_habil, n_days - 1, roll="forward",
                                  weekmask=WEEKMASK, holidays=FESTIVOS)
    else:
        # Día hábil → siguiente hábil
        siguiente_habil = np.busday_offset(date_part, 1, roll="forward",
                                           weekmask=WEEKMASK, holidays=FESTIVOS)
        limite = np.busday_offset(siguiente_habil, n_days - 1, roll="forward",
                                  weekmask=WEEKMASK, holidays=FESTIVOS)

    return datetime.combine(pd.to_datetime(str(limite)).date(), time_part)

# ------------------------------------------------------------
# FUNCIÓN: DÍAS HÁBILES ENTRE DOS FECHAS
# ------------------------------------------------------------
def business_days_between(start_dt, end_dt):
    if pd.isna(start_dt) or pd.isna(end_dt):
        return np.nan
    start_date = np.datetime64(start_dt.date() + timedelta(days=1))
    end_date = np.datetime64(end_dt.date())
    dias = np.busday_count(start_date, end_date, weekmask=WEEKMASK, holidays=FESTIVOS)
    if np.is_busday(end_date, weekmask=WEEKMASK, holidays=FESTIVOS) and end_date > start_date:
        dias += 1
    return int(dias)

# ------------------------------------------------------------
# CARGA DE DATOS
# ------------------------------------------------------------
df = pd.read_excel(ruta_input)
print(f"📂 Archivo cargado: {ruta_input.name} ({len(df)} registros)")

# ------------------------------------------------------------
# LIMPIEZA Y CONVERSIÓN DE FECHAS
# ------------------------------------------------------------
columnas_clave = ["PEDIDO", "FECHA_INICIO_ANS", "TIPO_DIRECCION", "ACTIVIDAD"]

for col in columnas_clave:
    if np.issubdtype(df[col].dtype, np.datetime64):
        df[col] = df[col].apply(lambda x: np.nan if pd.isna(x) else x)
    else:
        df[col] = df[col].apply(lambda x: np.nan if str(x).strip() == "" or str(x).upper() in ["NAN", "NONE", "NULL"] else x)

df["FECHA_INICIO_ANS"] = pd.to_datetime(df["FECHA_INICIO_ANS"], errors="coerce", dayfirst=True)

# ------------------------------------------------------------
# DÍAS PACTADOS
# ------------------------------------------------------------
DIAS_PACTADOS_MAP = {
    "ACREV":  {"URBANO": 4,  "RURAL": 4},
    "ALEGN":  {"URBANO": 7,  "RURAL": 10},
    "ALEGA":  {"URBANO": 7,  "RURAL": 10},
    "ACAMN":  {"URBANO": 7,  "RURAL": 10},
    "AMRTR":  {"URBANO": 7,  "RURAL": 10},
    "REEQU":  {"URBANO": 11, "RURAL": 11},
    "INPRE":  {"URBANO": 11, "RURAL": 11},
    "DIPRE":  {"URBANO": 11, "RURAL": 11},
    "ARTER":  {"URBANO": 5,  "RURAL": 8},
    "AEJDO":  {"URBANO": 5,  "RURAL": 8},
}

def dias_pactados(row):
    act = str(row.get("ACTIVIDAD", "")).strip().upper()
    tipo = str(row.get("TIPO_DIRECCION", "")).strip().upper()
    if act in DIAS_PACTADOS_MAP and tipo in DIAS_PACTADOS_MAP[act]:
        return DIAS_PACTADOS_MAP[act][tipo]
    return 0

df["DIAS_PACTADOS"] = df.apply(dias_pactados, axis=1)

# ------------------------------------------------------------
# FECHA LÍMITE ANS
# ------------------------------------------------------------
df["FECHA_LIMITE_ANS"] = df.apply(
    lambda r: add_business_days_keep_time(r["FECHA_INICIO_ANS"], r["DIAS_PACTADOS"]),
    axis=1
)

# ------------------------------------------------------------
# DÍAS TRANSCURRIDOS
# ------------------------------------------------------------
hoy = datetime.now()

def ajustar_hora(fecha_inicio):
    if pd.isna(fecha_inicio):
        return hoy
    return hoy.replace(hour=fecha_inicio.hour, minute=fecha_inicio.minute, second=fecha_inicio.second, microsecond=0)

def calcular_dias_transcurridos(row):
    fecha_ini = row["FECHA_INICIO_ANS"]
    if pd.isna(fecha_ini):
        return ""
    hoy_ref = ajustar_hora(fecha_ini)
    dias_habiles = business_days_between(fecha_ini, hoy_ref)
    hora_inicio = fecha_ini.strftime("%H:%M")
    return f"{dias_habiles} días {hora_inicio}"

df["DIAS_TRANSCURRIDOS"] = df.apply(calcular_dias_transcurridos, axis=1)

# ------------------------------------------------------------
# DÍAS RESTANTES (ajuste exacto incluyendo fin de semana y hora)
# ------------------------------------------------------------
def calcular_dias_restantes(row):
    fecha_lim = row["FECHA_LIMITE_ANS"]
    fecha_ini = row["FECHA_INICIO_ANS"]
    if pd.isna(fecha_lim) or pd.isna(fecha_ini):
        return ""

    hoy = datetime.now()
    hora_ref = fecha_ini.time()

    # Si ya venció
    if hoy > fecha_lim:
        return "VENCIDO"

    # Calcular días hábiles restantes sin sumar extra
    dias_habiles = np.busday_count(
        np.datetime64(hoy.date()),
        np.datetime64(fecha_lim.date()),
        weekmask=WEEKMASK,
        holidays=FESTIVOS
    )

    # ✅ Ajuste: si el siguiente día hábil es el mismo del límite, poner 1 día
    if dias_habiles == 0 and hoy.date() != fecha_lim.date():
        dias_habiles = 1

    # Si el día límite es hoy
    if hoy.date() == fecha_lim.date():
        if hoy.time() < fecha_lim.time():
            return f"0 días {fecha_ini.strftime('%H:%M')}"
        else:
            return "VENCIDO"

    # Si hoy es viernes y el vencimiento es lunes (fin de semana de por medio)
    # => contar solo el lunes como 1 día
    siguiente_habil = np.busday_offset(
        np.datetime64(hoy.date()), 1, roll="forward", weekmask=WEEKMASK, holidays=FESTIVOS
    )
    if siguiente_habil == np.datetime64(fecha_lim.date()):
        dias_habiles = 1

    return f"{dias_habiles} días {fecha_ini.strftime('%H:%M')}"

df["DIAS_RESTANTES"] = df.apply(calcular_dias_restantes, axis=1)

# ------------------------------------------------------------
# ESTADO
# ------------------------------------------------------------
def calcular_estado(row):
    valor = row["DIAS_RESTANTES"]
    if valor == "VENCIDO":
        return "VENCIDO"
    if isinstance(valor, str) and "días" in valor:
        try:
            dias = int(valor.split()[0])
            if dias == 0:
                return "ALERTA_0 Días"  # especial 0 días
            elif dias <= 2:
                return "ALERTA"
            return "A TIEMPO"
        except:
            return "SIN FECHA"
    return "SIN FECHA"

df["ESTADO"] = df.apply(calcular_estado, axis=1)

# ------------------------------------------------------------
# VERIFICAR SI EL ARCHIVO FENIX_ANS ESTÁ ABIERTO
# ------------------------------------------------------------
import os
import tkinter as tk
from tkinter import messagebox

def verificar_archivo_abierto(ruta):
    """Verifica si el archivo Excel está en uso por Excel u otro proceso."""
    if os.path.exists(ruta):
        try:
            with open(ruta, "a"):
                pass  # Si puede abrirse, no está bloqueado
        except PermissionError:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror(
                "🚫 Archivo bloqueado",
                "El Informe' está abierto en Excel.\n\n"
                "🔒 Cierra el archivo y vuelve a ejecutar el proceso."
            )
            print("⛔ Proceso detenido: el archivo está abierto.")
            exit()
# ------------------------------------------------------------
# 🔗 CRUCE CON GOOGLE SHEETS – FORMULARIO CONTROL ANS
# ------------------------------------------------------------
import gspread
from google.oauth2.service_account import Credentials

try:
    # Ruta al archivo de credenciales del proyecto (Service Account)
    cred_path = base_path / "Control_ANS" / "control-ans-elite-f4ea102db569.json"

    # Definir los permisos de acceso solo lectura
    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    creds = Credentials.from_service_account_file(cred_path, scopes=scopes)

    # Conexión con Google Sheets
    client = gspread.authorize(creds)

    # ✅ ID real de tu hoja "Formulario Control ANS"
    SHEET_ID = "1bPLGVVz50k6PlNp382isJrqtW_3IsrrhGW0UUlMf-bM"

    # Abrir la hoja
    sheet = client.open_by_key(SHEET_ID)
    # Buscar automáticamente la hoja que contiene "Form" o "Respuesta"
    sheet_names = [ws.title for ws in sheet.worksheets()]
    target_name = None
    for name in sheet_names:
        if "FORM" in name.upper() or "RESPUESTA" in name.upper():
            target_name = name
            break

    if not target_name:
        raise Exception(f"No se encontró ninguna pestaña válida. Hojas disponibles: {sheet_names}")

    worksheet = sheet.worksheet(target_name)
    print(f"📄 Hoja detectada automáticamente: {target_name}")


    # Leer todos los registros de la hoja activa
    data = worksheet.get_all_records()
    df_form = pd.DataFrame(data)
    df_form.rename(columns=lambda x: str(x).strip().upper(), inplace=True)

    # Normalizar nombres de columnas
    if "NÚMERO DEL PEDIDO" in df_form.columns:
        df_form.rename(columns={"NÚMERO DEL PEDIDO": "PEDIDO"}, inplace=True)
    if "ESTADO DEL PEDIDO" in df_form.columns:
        df_form.rename(columns={"ESTADO DEL PEDIDO": "FORMULARIO_FENIX"}, inplace=True)

    # Convertir PEDIDO a texto para evitar errores de cruce
    df["PEDIDO"] = df["PEDIDO"].astype(str)
    df_form["PEDIDO"] = df_form["PEDIDO"].astype(str)

    # Cruce (tipo LEFT JOIN)
    df = df.merge(df_form[["PEDIDO", "FORMULARIO_FENIX"]], on="PEDIDO", how="left")

    # Rellenar vacíos
    df["FORMULARIO_FENIX"] = df["FORMULARIO_FENIX"].fillna("SIN DATO")

    print("🔗 Cruce con formulario en Google Sheets completado correctamente.")
    print(f"📊 Registros leídos desde formulario: {len(df_form)}")
except Exception as e:
    print(f"⚠️ Error durante la conexión o cruce con Google Sheets: {e}")
# ------------------------------------------------------------
# EXPORTAR ARCHIVO
# ------------------------------------------------------------
verificar_archivo_abierto(ruta_output)  # 👈 ESTA LÍNEA ES CLAVE
# ------------------------------------------------------------
# 🔧 NORMALIZAR FECHAS PARA EVITAR DESFASES EN POWER BI
# ------------------------------------------------------------
# Se exportan como texto plano ISO (no tipo datetime)
# Así Power BI las lee exactamente igual sin conversión de zona ni AM/PM

df["FECHA_INICIO_ANS"] = df["FECHA_INICIO_ANS"].apply(
    lambda x: x.strftime("%Y-%m-%d %H:%M:%S") if pd.notnull(x) else ""
)
df["FECHA_LIMITE_ANS"] = df["FECHA_LIMITE_ANS"].apply(
    lambda x: x.strftime("%Y-%m-%d %H:%M:%S") if pd.notnull(x) else ""
)


ruta_output.parent.mkdir(exist_ok=True)
with pd.ExcelWriter(ruta_output, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="FENIX_ANS")
    resumen = df["ESTADO"].value_counts().reset_index()
    resumen.columns = ["ESTADO", "CANTIDAD"]
    resumen.to_excel(writer, index=False, sheet_name="RESUMEN")

print("✅ Cálculos ANS completados correctamente.")
print(f"📁 Archivo exportado: {ruta_output}")


# ------------------------------------------------------------
# FORMATO CONDICIONAL EN EXCEL
# ------------------------------------------------------------
wb = load_workbook(ruta_output)
ws = wb["FENIX_ANS"]
ultima_fila = ws.max_row
col_estado = "V"
rango = f"${col_estado}$2:${col_estado}${ultima_fila}"

# 🔴 VENCIDO
ws.conditional_formatting.add(
    rango,
    FormulaRule(formula=[f'${col_estado}2="VENCIDO"'],
    fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
)
# 🟠 ALERTA (0 días)
ws.conditional_formatting.add(
    rango,
    FormulaRule(formula=[f'${col_estado}2="ALERTA_0 Días"'],
    fill=PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"))
)
# 🟡 ALERTA (1 o 2 días)
ws.conditional_formatting.add(
    rango,
    FormulaRule(formula=[f'${col_estado}2="ALERTA"'],
    fill=PatternFill(start_color="FFF200", end_color="FFF200", fill_type="solid"))
)
# 🟢 A TIEMPO
ws.conditional_formatting.add(
    rango,
    FormulaRule(formula=[f'${col_estado}2="A TIEMPO"'],
    fill=PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"))
)

wb.save(ruta_output)
print("🎨 Formato condicional aplicado correctamente en la hoja FENIX_ANS.")

# ------------------------------------------------------------
# 🎨 FORMATO CONDICIONAL PARA COLUMNA 'FORMULARIO_FENIX'
# ------------------------------------------------------------
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Font

ws = wb["FENIX_ANS"]
ultima_fila = ws.max_row
col_form = "W"  # Columna FORMULARIO_FENIX
rango_form = f"${col_form}$2:${col_form}${ultima_fila}"

# 🟢 Verde → "Ejecutado en Campo"
ws.conditional_formatting.add(
    rango_form,
    FormulaRule(formula=[f'EXACTO(${col_form}2,"Ejecutado en Campo")'],
                fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                font=Font(color="006100"))
)

# 🔴 Rojo → "Pendiente" o "En Proceso"
ws.conditional_formatting.add(
    rango_form,
    FormulaRule(formula=[f'O(EXACTO(${col_form}2,"Pendiente"),EXACTO(${col_form}2,"En Proceso"))'],
                fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                font=Font(color="9C0006"))
)

# 🟠 Naranja → "En Ejecución" o "Revisión"
ws.conditional_formatting.add(
    rango_form,
    FormulaRule(formula=[f'O(EXACTO(${col_form}2,"En Ejecución"),EXACTO(${col_form}2,"Revisión"))'],
                fill=PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
                font=Font(color="7F6000"))
)

# ⚪ Gris claro → "SIN DATO"
ws.conditional_formatting.add(
    rango_form,
    FormulaRule(formula=[f'EXACTO(${col_form}2,"SIN DATO")'],
                fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
                font=Font(color="404040"))
)

# 💾 Guardar formato
wb.save(ruta_output)
print("🎨 Formato condicional aplicado correctamente en la columna FORMULARIO_FENIX.")

# ------------------------------------------------------------
# 💄 FORMATO VISUAL DE TABLA ESTRUCTURADA
# ------------------------------------------------------------
from openpyxl.worksheet.table import Table, TableStyleInfo  
from openpyxl.styles import Alignment

ws = wb["FENIX_ANS"]
ultima_fila = ws.max_row
ultima_col = ws.max_column
ultima_col_letra = ws.cell(row=1, column=ultima_col).column_letter

# Definir rango completo de la tabla
rango_tabla = f"A1:{ultima_col_letra}{ultima_fila}"

# Crear tabla estructurada si no existe
tabla = Table(displayName="FENIX_ANS_TABLA", ref=rango_tabla)

# Estilo sobrio (gris claro sin colores fuertes)
estilo = TableStyleInfo(
    name="TableStyleMedium2",  # azul corporativo con filtros
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
tabla.tableStyleInfo = estilo
ws.add_table(tabla)

# ------------------------------------------------------------
# 💡 Ajustes de formato visual
# ------------------------------------------------------------

# Quitar cuadrículas (solo visual, no elimina datos)
ws.sheet_view.showGridLines = False

# Ajustar ancho de columnas automáticamente
for col in ws.columns:
    max_len = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_len + 2

# Centrar columnas TELEFONO_CONTACTO y CELULAR_CONTACTO
for col_name in ["I", "J"]:  # ajusta si cambia la posición
    for cell in ws[col_name]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save(ruta_output)
print("💄 Formato visual de tabla estructurada aplicado correctamente.")

# ------------------------------------------------------------
# 📋 HOJA ADICIONAL: CONFIG_DIAS_PACTADOS
# ------------------------------------------------------------
from openpyxl.utils import get_column_letter

# Volver a abrir el archivo recién guardado
wb = load_workbook(ruta_output)

# Si ya existe la hoja, eliminarla para actualizarla
if "CONFIG_DIAS_PACTADOS" in wb.sheetnames:
    del wb["CONFIG_DIAS_PACTADOS"]

ws_conf = wb.create_sheet("CONFIG_DIAS_PACTADOS")

# Encabezados
headers = ["Actividad", "Descripción", "Días pactados Urbanos", "Días pactados Rurales"]
ws_conf.append(headers)

# Datos fijos según tu tabla
datos_dias = [
    ["ACREV", "PUNTOS DE CONEXIÓN", 4, 4],
    ["ALEGN", "LEGALIZACION", 7, 10],
    ["ALEGA", "LEGALIZACION", 7, 10],
    ["ACAMN", "REFORMA", 7, 10],
    ["AMRTR", "MOVIMIENTO REDES", 7, 10],
    ["REEQU", "TRABAJO ENERGÍA PREPAGO", 11, 11],
    ["INPRE", "INSTALACIÓN", 11, 11],
    ["DIPRE", "DESINSTALAR", 11, 11],
    ["ARTER", "REPLANTEO", 5, 8],
    ["AEJDO", "EJECUCIÓN", 5, 8],
]
for fila in datos_dias:
    ws_conf.append(fila)

# ------------------------------------------------------------
# 💄 FORMATO VISUAL
# ------------------------------------------------------------
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Bordes finos
thin_border = Border(
    left=Side(style='thin', color="BFBFBF"),
    right=Side(style='thin', color="BFBFBF"),
    top=Side(style='thin', color="BFBFBF"),
    bottom=Side(style='thin', color="BFBFBF")
)

# Encabezados en negrita, centrados, con fondo suave
for cell in ws_conf[1]:
    cell.font = Font(bold=True, color="000000")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    cell.border = thin_border

# Bordes y alineación general
for row in ws_conf.iter_rows(min_row=2, max_row=ws_conf.max_row, min_col=1, max_col=4):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

# Ajuste de ancho de columnas
for col in ws_conf.columns:
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws_conf.column_dimensions[col[0].column_letter].width = max_len + 2

# ------------------------------------------------------------
# Guardar con reintento (por bloqueo de OneDrive)
# ------------------------------------------------------------
import time
for intento in range(3):
    try:
        wb.save(ruta_output)
        print("💾 Archivo guardado correctamente con hoja CONFIG_DIAS_PACTADOS.")
        break
    except PermissionError:
        print("⚠️ Archivo temporalmente bloqueado. Reintentando...")
        time.sleep(2)
else:
    print("❌ No se pudo guardar el archivo. Cierra Excel o pausa OneDrive e inténtalo de nuevo.")

from datetime import datetime
from openpyxl import load_workbook

# # ------------------------------------------------------------
# # 📋 HOJA META_INFO - Información del proceso
# # ------------------------------------------------------------
wb = load_workbook(ruta_output)

# Si existe, eliminar para actualizar
if "META_INFO" in wb.sheetnames:
    del wb["META_INFO"]

ws_meta = wb.create_sheet("META_INFO")

ws_meta["A1"] = "Fuente de datos"
ws_meta["B1"] = "FENIX"

ws_meta["A2"] = "Fecha procesamiento Python"
ws_meta["B2"] = datetime.now().strftime("%d/%m/%Y %I:%M %p")

ws_meta["A3"] = "Archivo origen"
ws_meta["B3"] = "pendientes_FENIX.csv"

wb.save(ruta_output)
print("🧾 Hoja META_INFO agregada con fecha y hora del procesamiento.")


